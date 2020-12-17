#!/usr/bin/env python3
# vim: set fileencoding=utf-8 :
#from __future__ import print_function
from __future__ import unicode_literals
import sys
from openpyxl import Workbook, load_workbook
import xml.etree.ElementTree as ET
# from the modelica code, we will figure out what parameters are available
# for use to set...

class Tree(object):
	""" Tree structure, for storing available parameters.
	
	We will use this tree structure to save the names of the allowable 
	parameters from our Modelica code. Then we will read the information of the 		
	parameters from an Excel file. The information includes:
	1) nominal value
	2) unit
	3) type, i.e. 0-certain constant, 1-uncertain constant, 2-variable
	4) distribution
	5) boundary 1&2
 	The information will be passed to st_DAKOTA for sampling, and the sampled values 
	will be passed to Modelica when initialising our model. Alternatively, the nominal 
	values can be passed directly to Modelica when initilising our model.

	>>> T = Tree()
	>>> r = T.add_child('SM')	
	>>> r.add_value('type',2)
	>>> r.add_value('nominal',2.5)
	>>> r.add_value('unit',None)


	>>> T
	/
	  SM
		nominal = 2.5
		type = 2
		unit = None

	>>> T.insert('SM.distribution','uniform')
	>>> T.insert('SM.boundary1',1)
	>>> T.insert('SM.boundary2',5)

	>>> T
	/
	  SM
		boundary1 = 1
		boundary2 = 5
		distribution = 'uniform'
		nominal = 2.5
		type = 2
		unit = None

	>>> assert T.exists('SM.distribution')
	>>> assert not T.exists('field.foo')
	>>> assert not T.exists('foo.bar')

	>>> T.get('SM.nominal')
	2.5
	>>> T.update('SM.nominal',6)
	>>> T
	/
	  SM
		boundary1 = 1
		boundary2 = 5
		distribution = 'uniform'
		nominal = 6
		type = 2
		unit = None

	"""
	def __init__(self):
		self.children = {}
	def add_value(self,name,value,replace=False):
		assert replace is True or name not in self.children, "Child '%s' already exists"%(name,)
		self.children[name] = ValueNode(name,value)

	def add_child(self,name,replace=False):	
		assert replace is True or name not in self.children, "Child '%s' already exists"%(name,)
		self.children[name] = Tree()
		return self.children[name]

	def __repr__(self,prefix="",name="/"):
		if len(self.children) == 0:
			s = prefix + name + " {empty}\n"
		else:
			s = prefix + name + "\n"
			nextprefix = prefix + "  "
			for c in sorted(self.children):
				s += self.children[c].__repr__(nextprefix,c)
		if prefix == "":
			s = s.rstrip()
		return s

	def insert(self,ref,value,replace=False):
		#print("inserting",ref)
		if isinstance(ref,str):
			ref = ref.split(".")
		if ref[0] not in  self.children:
			if len(ref)==1:
				#print("adding '%s' with value"%(ref[0],),value)
				self.add_value(ref[0],value,replace=replace)
			else:
				self.add_child(ref[0]).insert(ref[1:],value,replace=replace)
		else:
			self.children[ref[0]].insert(ref[1:],value,replace=replace)

	def exists(self,ref):
		try:
			self.get(ref,raise_missing=True)
		except:
			return False
		return True

	def get(self,ref,raise_missing=False):
		#print("checking for",ref)
		
		if sys.version_info[0] >= 3:
			tp= str
		else:
			tp= unicode

		if isinstance(ref,str) or isinstance(ref,tp):
			ref = ref.split(".")

		if ref[0] not in self.children:
			#print("missing '%s'" % ref[0])
			if raise_missing:
				raise IndexError
			else:
				#print(ref[0])
				#print('not in the tree')
				return None
		if len(ref) == 1:
			if isinstance(self.children[ref[0]],Tree):
				return True
			else:
				return self.children[ref[0]].value
		else:
			assert isinstance(self.children[ref[0]],Tree),"should be tree node"
			return self.children[ref[0]].get(ref[1:],raise_missing)

	def update(self,ref,value):
		if isinstance(ref,str):
			ref = ref.split(".")
		assert ref[0] in self.children, "Child '%s' not found" % (ref[0],)
		if isinstance(self.children[ref[0]],ValueNode):
			self.children[ref[0]].value = value
		else:
			self.children[ref[0]].update(ref[1:],value)

	def load_xml(self, input_xml):
		"""
		input_xml: str, directory of the input xml file
		"""
		self.xml_tree = ET.parse(input_xml)	
		self.xml_root=self.xml_tree.getroot()
		
		isChangable=self.xml_root.findall("*ScalarVariable[@isValueChangeable='true']") # a list

		for par in isChangable:
			name=par.attrib["name"]
			name=name.split(".")
			if len(name)==1: # direct parameters in the main model (system-level)	
				name=name[0]
				start=self.xml_root.find('*ScalarVariable[@name=\''+name+'\']/*[@start]')
				#note: some of the variable does not have a 'start' value
				if start!=None:
					value=start.attrib['start']
					if par.attrib.has_key("description"):
						description=par.attrib["description"]
						classification, description=self.get_classification(description)
					else:
						description='-'
						classification='-'

					if start.attrib.has_key('unit'):
						unit=start.attrib['unit']
					else:
						unit='-'
					r=self.add_child(name, replace=True)
					r.add_value('nominal', value, replace=True)
					r.add_value('unit', unit, replace=True)
					r.add_value('description', description, replace=True)
					r.add_value('classification', classification, replace=True)


	def write_xml(self, output_xml):
		"""
		Export the nominal values in the tree to the xml file

		Argument:
		output_xml: str, directory of the updated (output) xml file
		"""		
		names=self.children.keys()
		for n in names:  
			v=self.get(n+'.nominal')
			u=self.get(n+'.unit')
			#d=self.get(n+'.description')

			find=self.xml_root.find('*ScalarVariable[@name=\''+n+'\']')
			if find!=None:
				changable=find.attrib['isValueChangeable']
				if changable=='true':
					self.xml_root.find('*ScalarVariable[@name=\''+n+'\']/*[@start]').attrib['start'] = str(v)
					self.xml_root.find('*ScalarVariable[@name=\''+n+'\']/*[@start]').attrib['unit'] = str(u)
					#self.xml_root.find('*ScalarVariable[@name=\''+n+'\']').attrib['description'] = str(d)

		self.xml_tree.write(output_xml)


	def update_xml(self, xmlfile):
		"""
		Update the values of changable parameters in the original xml file  
		with the nominal values in the tree

		Argument:
		xmlfile: str, directory of the xml file to be updated

		"""		

		xml_tree = ET.parse(xmlfile)	
		xml_root=xml_tree.getroot()

		names=self.children.keys()
		for n in names:  
			v=self.get(n+'.nominal')
			u=self.get(n+'.unit')
			#d=self.get(n+'.description')

			find=xml_root.find('*ScalarVariable[@name=\''+n+'\']')
			if find!=None:
				changable=find.attrib['isValueChangeable']
				if changable=='true':
					xml_root.find('*ScalarVariable[@name=\''+n+'\']/*[@start]').attrib['start'] = str(v)
					xml_root.find('*ScalarVariable[@name=\''+n+'\']/*[@start]').attrib['unit'] = str(u)
					#xml_root.find('*ScalarVariable[@name=\''+n+'\']').attrib['description'] = str(d)

		xml_tree.write(xmlfile)


	def filter_type(self, pmtype):
		"""
		Filter the specific type of parameters

		Argument:
		pmtype: int, 
				0-certain constant
				1-uncertain constant (for sensitivity analysis)
				2-variable (for optimisation or parametric study)

		Return:
		pmlist: list, a list of parameters that are pmtype
		"""
		pmlist=[]
		for n in self.children.keys():
			t=self.get(n+'.type')
			if t==pmtype:
				pmlist.append(n)

		return pmlist

	def filter_dist(self, dist, inlist):
		"""
		Filter a specific distribution

		Argument:
			dist: str, 
				'uniform': uniform distribution
				'norm': normal distribution
				'pert': pert-beta distribution
			inlist: a list of string, names of the parameters for checking

		Return:
			outlist: a list of string, 
				names of the parameters that have the specified distribution
		"""
		outlist=[]
		for n in inlist:
			d=self.get(n+'.distribution')
			if d==dist:
				outlist.append(n)
		return outlist


	def get_classification(self, description):
		'''
		description, str, the description of a parameter 
		'''
		if '[SYS]' in description:
			c='System level'
			description=description.replace('[SYS]','')

		elif '[H&T]' in description:
			c='Heliostat field and tower'
			description=description.replace('[H&T]','')

		elif '[RCV]' in description:
			c='Receiver'
			description=description.replace('[RCV]','')

		elif '[ST]' in description:
			c='Storage'
			description=description.replace('[ST]','')

		elif 'PB' in description:
			c='Power block'
			description=description.replace('[PB]','')

		elif '[CTRL]' in description:
			c='Control system'
			description=description.replace('[CTRL]','')

		elif '[FN]' in description:
			c='Finance'
			description=description.replace('[FN]','')
		else:
			c='NEC' # not elsewhere classified

		return c, description




class ValueNode(object):
	def __init__(self,name,value):
		self.value = value
	def __repr__(self,prefix="",name="{unnamed}"):
		return prefix + name + " = " + repr(self.value) + "\n"


def load_values_from_excel(filename,tree):
	"""
	Load allowable values from an Excel spreadsheet. 
	The template of the Excel spreadsheet is in the examples folder.
		Column A: classification
		Column B: symbol, should correspond to the names in Modelica
		Column C: nominal value
		Column D: unit
		Column E: name (description)
		Column F: type (0-certain constant, 1-uncertain constant, 2-variable)
				  type 1 is for uncertainty analysis
				  type 2 is for optimisation or parametric study 
		Column G: distribution of sampling, e.g. normal, uniform, pert-beta
		Column H: boundary1 (lower boundary or standard deviation)
		Column I: boundary2 (upper boundary if needed)
		Column J: description
		Column K: source or reference

	>>> T = Tree()
	>>> T.insert('SM.nominal',None)
	>>> T.insert('SM.distribution',None)
	>>> T.insert('rcv_H',None)
	>>> load_values_from_excel('Reference_2_params.xlsx',T)
	/
	  SM
		boundary1 = None
		boundary2 = None
		distribution = None
		nominal = 2.5
		type = 2
		unit = None
	  eff_blk
		boundary1 = 0.05
		boundary2 = None
		distribution = 'normal'
		nominal = 0.3774
		type = 1
		unit = None
	  he_av_design
		boundary1 = 0.985
		boundary2 = 0.995
		distribution = 'uniform'
		nominal = 0.99
		type = 1
		unit = None
	  rec_fr
		boundary1 = 0.01
		boundary2 = 0.09
		distribution = 'pert'
		nominal = 0.05
		type = 1
		unit = None
	  rcv_H
	"""
	wb = load_workbook(filename, data_only=True) # data_only=True, read value only, not he fomular

	for wsname in wb.sheetnames:
		ws = wb[wsname]
		for col in ws.iter_cols(min_col=2,max_col=2,min_row=ws.min_row, max_row=ws.max_row): # look at cells in column 
			#print("col =",col)
			for c in col:
				if c.value is not None: # name is there
					study = ws.cell(c.row,6).value
					if study in [0,1,2]:
						r=tree.add_child(c.value,replace=True)
						v = ws.cell(c.row,3).value
						assert v is not None, "No value next to label '%s'"%(c.value,)
						r.add_value('type', study,replace=True)
						r.add_value('nominal', v,replace=True)
						#r.add_value('unit', ws.cell(c.row,4).value,replace=True)
						r.add_value('distribution', ws.cell(c.row,7).value,replace=True)
						r.add_value('boundary1', ws.cell(c.row,8).value,replace=True)	
						r.add_value('boundary2', ws.cell(c.row,9).value,replace=True)	
						#r.add_value('description', ws.cell(c.row,10).value,replace=True)						

	return tree


def export_values_to_excel(filename, tree, inputxml=None):
	"""
	Export the parameters in the Tree to an excel spreadsheet (SolarTherm template)
		Column A: classification
		Column B: symbol, should correspond to the names in Modelica
		Column C: nominal value
		Column D: unit
		Column E: name (description)
		Column F: type (0-certain constant, 1-uncertain constant, 2-variable)
				  type 1 is for uncertainty analysis
				  type 2 is for optimisation or parametric study 
		Column G: distribution of sampling, e.g. normal, uniform, pert-beta
		Column H: boundary1 (lower boundary or standard deviation)
		Column I: boundary2 (upper boundary if needed)
		Column J: description
		Column K: source or reference

	Arguments:
		filename (str): the name and directory of the excel spreadsheet
		tree (class): the params.Tree()

	>>> T = Tree()
	>>> T.load_xml(inputxml) (optional)
	>>> export_values_to_excel('Export_params.xlsx',T)
	or
	>>> export_values_to_excel('Export_params.xlsx',T, inputxml)
	"""
	if inputxml!=None:
		tree.load_xml(inputxml)

	book=Workbook()
	sheet=book.active
	sheet['A2']='Classification'
	sheet['B2']='Symbol'
	sheet['C2']='Nominal'
	sheet['D2']='Unit'
	sheet['E2']='Name'
	sheet['F2']='Type'
	sheet['G2']='Distribution'
	sheet['H2']='B1(LB or SD)'
	sheet['I2']='B2 (UB)'
	sheet['J2']='Description'
	sheet['K2']='Source'
		
	names=tree.children.keys()
	length=len(names)

	for i in range(length):

		name=names[i]
		value=tree.get(name+'.nominal')
		unit=tree.get(name+'.unit')	
		description=tree.get(name+'.description')
		classification=tree.get(name+'.classification')			

		class_cell='A%s'%(i+3)
		symbol_cell='B%s'%(i+3)
		value_cell='C%s'%(i+3)		
		unit_cell='D%s'%(i+3)
		info_cell='K%s'%(i+3)
		description_cell='J%s'%(i+3)
		sheet[class_cell]=classification
		sheet[symbol_cell]=name
		sheet[value_cell]=value
		sheet[unit_cell]=unit
		sheet[description_cell]=description
		sheet[info_cell]='exported from xml file : %s'%inputxml	

	book.save(filename)



# TODO need to revise the tree structure
# there is only one tab in the spreadsheet

# TODO need a function to filter different type of parameters, e.g. 0, 1, 2

# TODO a new class to specify different distributions for uncertainty analysis
# e.g. uniform, normal, pert-beta distributions


if __name__ == "__main__":
	excel='/home/yewang/solartherm-master/examples/Reference_2_params.xlsx'
	input_xml='/media/yewang/Data/svn_gen3p3/system-modelling/research/sensitivity-analysis-DAKOTA/Reference_2_init.xml'
	output_xml='test_output.xml'

	T = Tree()
	T.load_xml(input_xml)
	print('original sm',T.get('SM.nominal'))
	load_values_from_excel(excel,T)
	print('')
	print('updated sm',T.get('SM.nominal'))
	print('eff_blk',T.get('eff_blk.type'))
	#T.write_xml(output_xml)
	pmlist=T.filter_type(1)
	print(pmlist)




